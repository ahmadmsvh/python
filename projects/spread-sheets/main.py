from user import User
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]

inv_file.create_sheet("Sheet2")
my_sheet = inv_file["Sheet2"]

product_no = 1
inventory = 2
price = 3
supplier = 4

products_per_supplier = {}
value_per_supplier = {}
product_under_ten = {}

for product_row in range(2,product_list.max_row + 1):
    supplier_name = product_list.cell(product_row , supplier).value
    value_per_inventory = product_list.cell(product_row ,  inventory).value * product_list.cell(product_row ,  price).value
    
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    
    else:
        products_per_supplier[supplier_name] = 1
        
    if supplier_name in value_per_supplier:
        value_per_supplier[supplier_name] += value_per_inventory
    else:
        value_per_supplier[supplier_name] = value_per_inventory
        
    product_inventory = product_list.cell(product_row , inventory).value
    product_number = product_list.cell(product_row,product_no).value
    if product_inventory < 10 :
        product_under_ten[product_number] = product_inventory
        
    product_price = product_list.cell(product_row , price).value
    product_total_value = product_inventory * product_price
    product_list.cell(product_row , 5).value = product_total_value
    my_sheet.cell(product_row , 5).value = product_total_value
    
print(products_per_supplier)        
print(product_under_ten)
product_list.cell(1,5).value = "Total price"    
my_sheet.cell(1,5).value = "Total price"    
my_sheet.cell(1,1).value = " "     
    
inv_file.save("inventory_new.xlsx")

ahmad = User("ahmadmsvh@gmail.com","ahmad mousavi","master developer","123456")
ahmad.get_user_info()

ali = User("ali.b@gamil.com","ali babaei","katoonibasketball","bb1234")
ali.get_user_info()